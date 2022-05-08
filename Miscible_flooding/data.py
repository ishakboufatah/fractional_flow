# -*- coding: utf-8 -*-
"""
Created on Mon Aug 30 17:54:52 2021

@author: ishak
"""

import numpy as np
import openpyxl  as xl
ws=xl.load_workbook('OIL.xlsx',data_only=True)
SCAL= ws['SCAL'] 

Kr=xl.load_workbook('Relative Permeability Results Dead-Oil with Injection Water.xlsx',data_only=True)
SCAL2=Kr['Gas-Oil OMJ-323 Sample 4']
OMG_832_3=Kr['Gas-Oil OMG-832 Sample 3']
OMG_832_14=Kr['Gas-Oil OMG-832 Sample 14']
def get_value(ws,a):
    result=[]
    for cell in ws[a:]:
        result.append(cell.value)
    return result

########################################################################
#---------------------relative permeability data from ECLIPSE------------------------
 
Sw=get_value(SCAL['B'],1)     ;Sg=get_value(SCAL['H'],1) 
Krw=get_value(SCAL['C'],1)    ;Krg=get_value(SCAL['I'],1)
Krow=get_value(SCAL['D'],1)   ;Krog=get_value(SCAL['J'],1)

########################################################################



########################################################################
#---------------------relative permeability data from EXPIRIMRNT------------------------

#--------------------- Gas-Oil OMJ-323 Sample 4---------------------------
SG=get_value(SCAL2['D'],10) 
KrG=get_value(SCAL2['F'],10)
KrO=get_value(SCAL2['E'],10)

#print(SG)

#--------------------- Gas-Oil OMG-832 Sample 3---------------------------
SG_OMG_832_3=get_value(OMG_832_3['D'],10) 
KrG_OMG_832_3=get_value(OMG_832_3['F'],10)
KrO_OMG_832_3=get_value(OMG_832_3['E'],10)

#print(SG_OMG_832_3)


#--------------------- Gas-Oil OMG-832 Sample 14---------------------------
SG_OMG_832_14=get_value(OMG_832_14['D'],10) 
KrG_OMG_832_14=get_value(OMG_832_14['F'],10)
KrO_OMG_832_14=get_value(OMG_832_14['E'],10)

#print(SG_OMG_832_14)
########################################################################
#print(Sw)
